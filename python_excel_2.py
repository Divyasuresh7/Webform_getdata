from flask import Flask, render_template, request,session,redirect,url_for
from openpyxl import load_workbook,Workbook
import datetime
from forms import SignInForm

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret-key'
EXCEL_FILE_NAME = 'form_data.xlsx'
x = datetime.datetime.now()

@app.route('/home')
def home():
    return render_template('base.html')

@app.route('/signup',methods=['GET','POST'])
def signup():
    return render_template('new_web_form.html')

users = {'divya':'password','suresh':'password'}

@app.route('/signin',methods=['GET','POST'])
def signin():
    form = SignInForm()
    if form.is_submitted():
        result = request.form()
        return render_template('user.html',result = result)
    return render_template('signin.html',form=form)

@app.route('/submit', methods=['POST'])
def submit():
    # To get data from the form
    date_1 = request.form['Date']
    emp_name = request.form['emp_name']
    company_name = request.form['Company_name']
    customer_name = request.form['Customer_name']
    date_2 = request.form['Meeting_time']
    message = request.form['message']
    date_3 = request.form['Follow_up']

    # Load the existing excel or create a new one if it doesn't already exist
    try:
        workbook = load_workbook(EXCEL_FILE_NAME)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet or create a new one if it doesn't exist
    sheet = workbook.active if workbook.sheetnames else workbook.create_sheet()

    # Check if the sheet has column headers, if not, set them
    if not sheet['A1'].value:
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Employee Name'
        sheet['C1'] = 'Company Name'
        sheet['D1'] = 'Customer Name'
        sheet['E1'] = 'Meeting Time'
        sheet['F1'] = 'Agenda'
        sheet['G1'] = 'Follow Up'

    # Get the next available row
    row = sheet.max_row + 1

    # Write form data to Excel
    sheet.cell(row=row, column=1, value=date_1)
    sheet.cell(row=row, column=2, value=emp_name)
    sheet.cell(row=row, column=3, value=company_name)
    sheet.cell(row=row, column=4, value=customer_name)
    sheet.cell(row=row, column=5, value=date_2)
    sheet.cell(row=row, column=6, value=message)
    sheet.cell(row=row, column=7, value=date_3)

    # Save the workbook
    workbook.save(EXCEL_FILE_NAME)

    return 'Form data submitted successfully!'

if __name__ == '__main__':
    app.run()
