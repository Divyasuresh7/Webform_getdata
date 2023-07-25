from flask import Flask, render_template, request,redirect,url_for
from openpyxl import load_workbook,Workbook
import datetime
import pandas as pd
import csv
from django.shortcuts import render,redirect
import calendar 
from calendar import HTMLCalendar

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret-key'
EXCEL_FILE_NAME = 'form_data.xlsx'
x = datetime.datetime.now()

@app.route('/home')
def home():
    return render_template('base.html')

@app.route('/signin',methods=['GET','POST'])
def signin():
    return render_template('signin.html')

@app.route('/signup', methods=['POST'])
def submited():
    # To get data from the form
    result1 = request.form['username']
    result2 = request.form['password']
    global result
    result = result1

    def validate_credentials(username, password):
        with open('username_password.csv', 'r') as file:
            reader = csv.reader(file)
            for row in reader:
                stored_username = row[0]
                stored_password = row[1]
                if username == stored_username and password == stored_password:
                    return True
        return False
    if validate_credentials(result1,result2):
        return render_template('new_web_form.html',empl_name=result)
    else:
        return "Invalid username or password."
    

@app.route('/submit', methods=['POST'])
def submit():
    # To get data from the form
    date_1 = request.form['Date']
    emp_name = request.form['emp_name']
    company_name = request.form['Company_name']
    customer_name = request.form['Customer_name']
    date_2 = request.form['Meeting_time']
    message = request.form['message']
    date_3 = request.form['Follow_up']

    # Load the existing excel or create a new one if it doesn't already exist
    try:
        workbook = load_workbook(EXCEL_FILE_NAME)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet or create a new one if it doesn't exist
    sheet = workbook.active if workbook.sheetnames else workbook.create_sheet()

    # Check if the sheet has column headers, if not, set them
    if not sheet['A1'].value:
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Employee Name'
        sheet['C1'] = 'Company Name'
        sheet['D1'] = 'Customer Name'
        sheet['E1'] = 'Meeting Time'
        sheet['F1'] = 'Agenda'
        sheet['G1'] = 'Follow Up'

    # Get the next available row
    row = sheet.max_row + 1

    # Write form data to Excel
    sheet.cell(row=row, column=1, value=date_1)
    sheet.cell(row=row, column=2, value=emp_name)
    sheet.cell(row=row, column=3, value=company_name)
    sheet.cell(row=row, column=4, value=customer_name)
    sheet.cell(row=row, column=5, value=date_2)
    sheet.cell(row=row, column=6, value=message)
    sheet.cell(row=row, column=7, value=date_3)

    # Save the workbook
    workbook.save(EXCEL_FILE_NAME)

    return 'Form data submitted successfully!'

@app.route('/details_page',methods=['GET','POST'])
def details_option():
    df = pd.read_excel('Employee_details.xlsx',sheet_name='Sheet1')
    result = "Ridhima"
    # result = request.form['username']
    team_data = df[df['MANAGER'] == result]['NAME']
    team_data2 = team_data.to_string(index=False)
    team_data_list = team_data2.split()
    no_of_hrs =[]
    dff = pd.read_excel('form_data.xlsx') #data from form
    for i in team_data_list:
        emp_names = dff[dff['Employee Name']==i]
        team_data2 = emp_names.to_string(index=False)
        if i in team_data2:
            no_of_hrs.append(len(emp_names))
    def emp_check():
        if not team_data.empty:
            return team_data_list
        else:
            return "No employees found for this manager."
    names = emp_check()
    data={}
    for key in names:
        for value in no_of_hrs:
            data[key] = value
            no_of_hrs.remove(value)
            break
    return render_template('view_details.html',data=data)

@app.route('/view_inividual_details',methods=['GET','POST'])
def view_details(team_data2):
    return render_template('details.html',team_data2=team_data2)

@app.route('/details', methods=['POST'])
def view_emp_details():
    load_workbook(EXCEL_FILE_NAME)
    df = pd.read_excel(EXCEL_FILE_NAME)
    name = request.form['employee_name']
    team_data2 = df.loc[df['Employee Name'] == name]
    
    return view_details(team_data2)

if __name__ == '__main__':
    app.debug = True
    app.run()
